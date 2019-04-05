#coding: UTF-8
import requests
from bs4 import BeautifulSoup
import re
import openpyxl as excel
import urllib
import time

print('URLを入力してください')
url = input() 

data_list = []

r_top = requests.get(url) 
soup_top = BeautifulSoup(r_top.content, 'html.parser')
top_h1 = str(soup_top.find('h1'))
result = re.split('<|>|\(|\)', top_h1) 
area = result[2].replace('の人気美容院・美容室・ヘアサロン ', '') # エリア名
page = int(result[3].replace('1/', '')) # ページ数

for i in range(1): # ページネーションをたどる
    i = str(i+1)
    r_list = requests.get(url + 'PN' + i + '.html') 
    body = r_list.content
    soup = BeautifulSoup(body, 'html.parser')
    h3_list = soup.select('h3')
    for h3  in h3_list:
        for a in h3.select('a'):
            href = a.attrs['href']
            href = href.split('?')
            href = href[0]  #店舗URL
            shop_name = a.string #店舗名
            r_shop = requests.get(href) 
            soup_shop = BeautifulSoup(r_shop.content, 'html.parser')
            shop_adress = soup_shop.find('ul', class_='fs10')
            shop_adress = str(shop_adress.find('li')) 
            shop_adress = shop_adress.replace('<li>', '')
            shop_adress = shop_adress.replace('</li>', '') #サロン住所
            shop_link = str(soup_shop.find('div', class_='mT30 mB20'))
            shop_link = shop_link.count('<li>') #関連リンク数
            shop_tel = soup_shop.find('th', class_='w120')
            shop_tel = str(shop_tel.get_text())
            if shop_tel in '電話番号':
                time.sleep(2)
                tel_url = href + 'tel/'
                r_tel = requests.get(tel_url)
                soup_tel = BeautifulSoup(r_tel.content, 'html.parser')
                shop_tel = soup_tel.find('td', class_='fs16')
                shop_tel = shop_tel.get_text()
            else:
                shop_tel = 'なし'
            data_list.append([shop_name, shop_adress, shop_tel, shop_link, href])
            time.sleep(5)
    time.sleep(5)

# excel作成
# 新規ワークブックを作る
wb = excel.Workbook()
# アクティブなワークシートを得る
ws = wb.active
# A1のセルに値を設定
ws['A1'] = area
def write_list_2d(ws, data_list, start_row, start_col):
    for y, row in enumerate(data_list):
        for x, cell in enumerate(row):
            ws.cell(row=start_row + y,
                       column=start_col + x,
                       value=data_list[y][x])
write_list_2d(ws, data_list, 2, 1)

# ファイルを保存
wb.save(area + '.xlsx')
